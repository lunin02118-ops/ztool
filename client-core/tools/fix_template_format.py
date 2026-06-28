"""
Reformat bom_шаблон.xlsx for professional Russian appearance:
1. Rename sheet from 图明1 → Спецификация
2. Change all fonts from 宋 → Arial
3. Proper column widths
4. Proper row heights (especially for thumbnails)
5. Header formatting (bold, centered, borders)
6. Add stable Russian anchors for calculated ZTool columns
7. Update defined names to reference new sheet name
8. Fix alignment and number formats
"""
import os
import re
import sys
import zipfile
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.workbook.defined_name import DefinedName

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

# Resolve the template relative to the repo root so the script works no matter
# what the current working directory is.
_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
TEMPLATE = os.path.join(_ROOT, "Шаблоны спецификации", "bom_шаблон.xlsx")
OUTPUT = TEMPLATE

OLD_SHEET = "图明1"
NEW_SHEET = "Спецификация"
LEGACY_SHEET_NAMES = (OLD_SHEET, "图纸明细1")

# Re-run safe: also treat an already-renamed sheet as the source.
SONG = "宋"  # leftover vendor CJK font (SimSun) that must be purged everywhere

# Professional font
FONT_TITLE = Font(name="Arial", size=16, bold=True)
FONT_HEADER = Font(name="Arial", size=10, bold=True)
FONT_META = Font(name="Arial", size=10, bold=False)
FONT_DATA = Font(name="Arial", size=10, bold=False)

# Alignment
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)
ALIGN_LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Border
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

# Column widths (optimized for Russian text)
COL_WIDTHS = {
    "A": 5.5,    # № п/п
    "B": 12.0,   # Код материала
    "C": 18.0,   # Наименование
    "D": 16.0,   # Обозначение
    "E": 8.0,    # Версия
    "F": 7.0,    # Ед. изм.
    "G": 7.5,    # Кол-во
    "H": 10.0,   # Тип обработки
    "I": 18.0,   # Обработка поверхности
    "J": 8.0,    # Масса
    "K": 12.0,   # Примечание
    "L": 16.0,   # Имя файла
    "M": 12.0,   # Эскиз (for thumbnails)
    "N": 14.0,   # Материал
    "O": 35.0,   # Путь
    "P": 18.0,   # Габарит
}

DATA_COLS = 16

# Data row height
DATA_ROW_HEIGHT = 20.0
# Thumbnail row height (for modes with images)
THUMBNAIL_ROW_HEIGHT = 50.0


def main():
    wb = load_workbook(TEMPLATE)
    ws = wb.active

    # 1. Rename sheet
    ws.title = NEW_SHEET
    print(f"[1] Sheet renamed: {OLD_SHEET} → {NEW_SHEET}")

    # 1b. BULLETPROOF: purge the leftover CJK font (宋/SimSun) from EVERY cell in
    # the existing used range. ZTool's exporter clones the style of whatever
    # template row sits at the data anchor when inserting data rows, so a single
    # stray 宋 cell in that path makes exported data render in SimSun. We
    # normalise the whole used range to Arial, preserving size/bold/italic/color.
    # Bound the scan to the EXISTING used range so we don't materialise new
    # (empty, bordered) rows/cols and bloat the template.
    #
    # NOTE: this per-cell pass cannot touch the non-anchor cells of a merged
    # range (they are MergedCell shadows; openpyxl silently ignores font
    # assignment on them) nor leftover entries in the workbook font table.
    # The absolute purge is done after save() in purge_cjk_fonts_in_styles_xml().
    max_row, max_col = ws.max_row, ws.max_column
    purged = 0
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            fn = cell.font
            if fn.name and fn.name != "Arial":
                cell.font = Font(
                    name="Arial",
                    size=fn.size or 10,
                    bold=fn.bold,
                    italic=fn.italic,
                    color=fn.color,
                )
                purged += 1
    print(f"[1b] Purged non-Arial font from {purged} cells (宋 → Arial)")

    # 2. Update all defined names to reference new sheet
    names_to_update = {}
    for name, defn in list(wb.defined_names.items()):
        old_ref = defn.attr_text
        new_ref = old_ref
        for legacy_sheet in LEGACY_SHEET_NAMES:
            new_ref = new_ref.replace(legacy_sheet, NEW_SHEET)
        if new_ref != old_ref:
            names_to_update[name] = new_ref

    for name, new_ref in names_to_update.items():
        del wb.defined_names[name]
        new_defn = DefinedName(name=name, attr_text=new_ref)
        wb.defined_names.add(new_defn)
        print(f"  Updated name: {name} → {new_ref}")

    # 2b. Add/normalize release-specific calculated columns and aliases.
    # The original workbook has no column for ZTool's computed bounding-box
    # values. Appending it avoids shifting existing legacy anchors.
    ws["J6"] = "Масса"
    ws["P6"] = "Габарит"
    calculated_names = {
        "Номер": f"'{NEW_SHEET}'!$A$6",
        "Количество": f"'{NEW_SHEET}'!$G$6",
        "Путь": f"'{NEW_SHEET}'!$O$6",
        "Эскиз": f"'{NEW_SHEET}'!$M$6",
        "МассаЕдКг": f"'{NEW_SHEET}'!$J$6",
        "ГабаритныеРазмеры": f"'{NEW_SHEET}'!$P$6",
        "ЕдИзм": f"'{NEW_SHEET}'!$F$6",
    }
    for name, ref in calculated_names.items():
        if name in wb.defined_names:
            del wb.defined_names[name]
        wb.defined_names.add(DefinedName(name=name, attr_text=ref))
    print("[1c] Service/calculated anchors added: Номер, Количество, Путь, Эскиз, МассаЕдКг, ГабаритныеРазмеры, ЕдИзм")

    # 3. Set column widths
    for col, width in COL_WIDTHS.items():
        ws.column_dimensions[col].width = width
    print("[2] Column widths set")

    # 4. Format title row (row 1)
    for col in range(1, DATA_COLS + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = FONT_TITLE
        cell.alignment = ALIGN_CENTER
    ws.row_dimensions[1].height = 28.0
    print("[3] Title row formatted")

    # 5. Format metadata rows (rows 2-5)
    for row in range(2, 6):
        ws.row_dimensions[row].height = 16.0
        for col in range(1, DATA_COLS + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = FONT_META
            cell.alignment = ALIGN_LEFT
    print("[4] Metadata rows formatted")

    # 6. Format header row (row 6) — bold, centered, with borders and fill
    ws.row_dimensions[6].height = 30.0
    for col in range(1, DATA_COLS + 1):
        cell = ws.cell(row=6, column=col)
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
        cell.fill = HEADER_FILL
    print("[5] Header row formatted (bold, borders, fill)")

    # 7. Format data rows (7 onwards) — borders, proper font, alignment
    for row in range(7, ws.max_row + 1):
        ws.row_dimensions[row].height = DATA_ROW_HEIGHT
        for col in range(1, DATA_COLS + 1):  # noqa: data columns A..P carry borders
            cell = ws.cell(row=row, column=col)
            cell.font = FONT_DATA
            cell.border = THIN_BORDER
            if col == 1:  # № п/п — center
                cell.alignment = ALIGN_CENTER
            elif col == 7:  # Кол-во — center
                cell.alignment = ALIGN_CENTER
            elif col == 10:  # Масса — center
                cell.alignment = ALIGN_CENTER
            elif col == 15:  # Путь — left, no wrap
                cell.alignment = ALIGN_LEFT
            elif col == 16:  # Габарит — center-ish but keep wrap
                cell.alignment = ALIGN_LEFT_WRAP
            else:
                cell.alignment = ALIGN_LEFT_WRAP
    print(f"[6] Data rows 7-{ws.max_row} formatted")

    # 8. Set thumbnail column (M) properties for image binding
    # Images in ZTool are inserted with anchor to the cell.
    # Row height needs to accommodate the image (64x48 default).
    # We'll set a reasonable default that works both with and without images.
    # When ZTool exports with images, it may adjust row heights dynamically.
    # M column width = 12 chars ≈ 86px (enough for 64px image + padding)
    ws.column_dimensions["M"].width = 12.0
    print("[7] Thumbnail column M configured")

    # 9. Freeze panes at row 7 (header stays visible when scrolling)
    ws.freeze_panes = "A7"
    print("[8] Freeze panes set at A7")

    # 9b. Page setup: the vendor template printed only A1:K40 (cut off columns
    # L..O and most rows). Make the printout professional: landscape, fit all
    # data columns to one page wide, repeat the title+header rows on every page.
    ws.print_area = "A1:P75"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = "$1:$6"
    print("[9] Page setup: landscape, fit-to-width, print area A1:P75")

    # 10. Save
    wb.save(OUTPUT)

    # 11. Absolute font purge at the XML level. The per-cell pass [1b] cannot
    # reach merged-range shadow cells or stale entries left in the workbook
    # font table, so a few CJK (宋/SimSun, 华行/...) <font> definitions survive
    # and keep rendering in the wrong typeface. Rewrite xl/styles.xml so every
    # font whose name contains a CJK codepoint becomes Arial, and drop the
    # charset="134" (GB2312) hints that make Excel substitute a CJK fallback.
    n_fonts, n_charset = purge_cjk_fonts_in_styles_xml(OUTPUT)
    print(f"[11] styles.xml purge: {n_fonts} CJK font(s) -> Arial, "
          f"removed {n_charset} GB2312 charset hint(s)")

    print(f"\n[DONE] Saved: {OUTPUT}")
    print(f"  Sheet: {NEW_SHEET}")
    print(f"  Defined names preserved and updated")


def _has_cjk(s):
    """True if the string contains any CJK / fullwidth codepoint."""
    return any(ord(ch) >= 0x2E80 for ch in s)


def purge_cjk_fonts_in_styles_xml(path):
    """Rewrite xl/styles.xml in-place so no <font> uses a CJK typeface.

    Returns (num_font_names_replaced, num_charset_hints_removed).
    Only the <fonts> block is touched; cell references are untouched, so every
    cell that pointed at a CJK font now renders in Arial.
    """
    with zipfile.ZipFile(path, "r") as zin:
        items = {name: zin.read(name) for name in zin.namelist()}

    xml = items["xl/styles.xml"].decode("utf-8")
    fonts_match = re.search(r"(<fonts\b.*?</fonts>)", xml, re.S)
    if not fonts_match:
        return 0, 0
    fonts = fonts_match.group(1)

    replaced = 0

    def _repl_name(m):
        nonlocal replaced
        if _has_cjk(m.group(1)):
            replaced += 1
            return '<name val="Arial" />'
        return m.group(0)

    new_fonts = re.sub(r'<name val="([^"]*)" />', _repl_name, fonts)
    n_charset = new_fonts.count('charset val="134"')
    new_fonts = new_fonts.replace('<charset val="134" />', "")

    xml = xml[: fonts_match.start(1)] + new_fonts + xml[fonts_match.end(1):]
    items["xl/styles.xml"] = xml.encode("utf-8")

    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, data in items.items():
            zout.writestr(name, data)
    return replaced, n_charset


if __name__ == "__main__":
    main()
