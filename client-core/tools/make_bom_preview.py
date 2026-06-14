"""
Build a filled PREVIEW of the BOM template (bom_шаблон.xlsx) with realistic
Russian sample data, so the formatting (Arial font, borders, column widths,
header fill) can be reviewed in Excel WITHOUT running ZTool/SolidWorks.

This does NOT change the template. It loads the template, writes a handful of
sample rows into the data area (rows 7+), and saves a separate *_preview.xlsx.
The font/borders/widths are inherited verbatim from the template, so the
preview is an accurate picture of what an export will look like.

Usage:
    python client-core/tools/make_bom_preview.py
"""
import os
from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
TEMPLATE = os.path.join(ROOT, "Шаблоны спецификации", "bom_шаблон.xlsx")
PREVIEW = os.path.join(ROOT, "Шаблоны спецификации", "bom_шаблон_preview.xlsx")

# Columns (1-based) as anchored in the template header row 6:
# A№ B Код C Наимен. D Обознач. E Версия F Ед.изм G Кол-во H Тип обр.
# I Обр.пов. J Масса K Примеч. L Имя файла M Эскиз N Материал O Путь
SAMPLE = [
    # №, Код,        Наименование,           Обозначение,  Вер, Ед, Кол, Тип обр.,  Обр.пов., Масса, Примеч.,  Имя файла,        Материал,    Путь
    (1, "СБ-001",  "Платформа захвата",     "0614-А00.01", "A", "шт", 1, "Сварка",   "Окраска", 12.4, "",       "0614-A00-01.sldasm", "Сталь 09Г2С", r"D:\Проекты\0614\0614-A00-01.SLDASM"),
    (2, "Д-014",   "Корпус захвата",        "0614-А00.02", "A", "шт", 6, "Мех.обр.", "Без",     0.82, "",       "0614-A00-02.sldprt", "Сталь 45",    r"D:\Проекты\0614\0614-A00-02.SLDPRT"),
    (3, "Д-021",   "Палец",                 "0614-А00.03", "B", "шт", 3, "Мех.обр.", "Хром",    0.15, "Закалка","0614-A00-03.sldprt", "Сталь 40Х",   r"D:\Проекты\0614\0614-A00-03.SLDPRT"),
    (4, "П-100",   "Подшипник 6204",        "ГОСТ 8338-75","-", "шт", 9, "Покупное", "-",       0.11, "",       "6204.sldprt",        "—",           r"D:\Проекты\0614\Покупные\6204.SLDPRT"),
    (5, "П-205",   "Винт М6×20",            "ГОСТ 11738-84","-","шт",12, "Покупное", "Оцинк.",  0.01, "",       "vint-m6.sldprt",     "Сталь 8.8",   r"D:\Проекты\0614\Крепёж\vint-m6.SLDPRT"),
    (6, "Д-033",   "Кронштейн",             "0614-А00.06", "A", "шт", 2, "Листовая", "Окраска", 1.65, "",       "0614-A00-06.sldprt", "Сталь 3",     r"D:\Проекты\0614\0614-A00-06.SLDPRT"),
    (7, "Д-040",   "Пластина опорная",      "0614-А00.07", "A", "шт", 4, "Мех.обр.", "Без",     0.44, "",       "0614-A00-07.sldprt", "АМг6",        r"D:\Проекты\0614\0614-A00-07.SLDPRT"),
    (8, "Д-051",   "Втулка направляющая",   "0614-А00.08", "C", "шт", 6, "Мех.обр.", "Без",     0.07, "",       "0614-A00-08.sldprt", "Бронза БрАЖ", r"D:\Проекты\0614\0614-A00-08.SLDPRT"),
]

# data column index -> tuple field index
COL_MAP = {1: 0, 2: 1, 3: 2, 4: 3, 5: 4, 6: 5, 7: 6, 8: 7, 9: 8, 10: 9,
           11: 10, 12: 11, 14: 12, 15: 13}


def main():
    wb = load_workbook(TEMPLATE)
    ws = wb.active
    start = 7
    for i, rowdata in enumerate(SAMPLE):
        r = start + i
        for col, field in COL_MAP.items():
            ws.cell(row=r, column=col).value = rowdata[field]
    # fill the assembly title/meta so the header block isn't blank
    ws["A2"] = "Наименование изделия: Платформа захвата 0614-А00"
    ws["A3"] = "Модель изделия: 0614-A00.SLDASM"
    wb.save(PREVIEW)
    print("Saved preview:", PREVIEW)
    print("Rows filled:", len(SAMPLE), "(rows 7..%d)" % (start + len(SAMPLE) - 1))


if __name__ == "__main__":
    main()
