#!/usr/bin/env python3
"""
validate_bom_exports.py — Validates exported BOM .xlsx files against expected
behavior for each of the 8 export modes.

Usage:
    python validate_bom_exports.py <export_directory>

The script looks for .xlsx files in the directory and validates:
- Row count (data rows present)
- Service columns: № п/п (A), Кол-во (G), Путь (O)
- Property columns: C, D, E, H, I, J, N (if model has matching propnames)
- Mode-specific behavior (type 0/1/2/3, images, filters)

Requires: openpyxl (pip install openpyxl)
"""
import sys
import os
import re
from pathlib import Path

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


# Mode definitions matching ZTool.settings bomsettings order
MODES = [
    {"id": 1, "name": "Экспорт сводной спецификации",
     "type": 0, "image": False, "filter": None,
     "desc": "Flat summary BOM, duplicates merged, Qty = total count"},
    {"id": 2, "name": "Экспорт иерархической спецификации",
     "type": 1, "image": False, "filter": None,
     "desc": "Indented BOM showing assembly hierarchy"},
    {"id": 3, "name": "Экспорт спецификации верхнего уровня",
     "type": 2, "image": False, "filter": None,
     "desc": "Top-level only (direct children of root assembly)"},
    {"id": 4, "name": "Экспорт сводной спецификации деталей",
     "type": 3, "image": False, "filter": None,
     "desc": "Parts only (no sub-assemblies)"},
    {"id": 5, "name": "Экспорт сводной спецификации (с эскизами)",
     "type": 0, "image": True, "filter": None,
     "desc": "Same as mode 1 + thumbnail images in column M"},
    {"id": 6, "name": "Экспорт иерархической спецификации (с эскизами)",
     "type": 1, "image": True, "filter": None,
     "desc": "Same as mode 2 + thumbnail images in column M"},
    {"id": 7, "name": "Обрабатываемые детали",
     "type": 0, "image": False, "filter": "Обрабатываемые детали",
     "desc": "Filtered: only machined/manufactured parts"},
    {"id": 8, "name": "Покупные изделия",
     "type": 0, "image": False, "filter": "Покупные изделия",
     "desc": "Filtered: only purchased/standard parts"},
]

# Column mapping (1-indexed): A=1, B=2, ..., O=15
COL_NUM = 1      # A - № п/п (service)
COL_QTY = 7      # G - Кол-во (service)
COL_PATH = 15    # O - Путь (service)
COL_IMAGE = 13   # M - Эскиз (service)
# Property columns (C=3, D=4, E=5, H=8, I=9, J=10, N=14)
PROP_COLS = [3, 4, 5, 8, 9, 10, 14]

DATA_START_ROW = 7  # Data starts at row 7 (row 6 is header anchor row)


def analyze_xlsx(filepath):
    """Analyze a single .xlsx export file."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    results = {
        "file": os.path.basename(filepath),
        "total_rows": 0,
        "num_filled": 0,
        "qty_filled": 0,
        "path_filled": 0,
        "prop_filled": 0,
        "has_images": False,
        "max_row": ws.max_row,
        "issues": [],
    }

    if ws.max_row < DATA_START_ROW:
        results["issues"].append("No data rows (max_row < 7)")
        return results

    # Determine the ACTUAL last data row: scan from the bottom for the last row
    # that has ANY meaningful content in service/property columns. This avoids
    # counting formatted-but-empty rows (template pre-formats rows down to 75).
    # A row counts as "data" if № (A), Кол-во (G), Путь (O), or any property
    # column has a non-empty value.
    significant_cols = [COL_NUM, COL_QTY, COL_PATH] + PROP_COLS
    last_data_row = DATA_START_ROW - 1
    for row in range(DATA_START_ROW, ws.max_row + 1):
        for col in significant_cols:
            v = ws.cell(row=row, column=col).value
            if v is not None and str(v).strip():
                last_data_row = row
                break

    if last_data_row < DATA_START_ROW:
        results["total_rows"] = 0
        results["issues"].append("No data rows (all empty below header)")
        return results

    data_rows = last_data_row - DATA_START_ROW + 1
    results["total_rows"] = data_rows
    results["last_data_row"] = last_data_row

    num_count = 0
    qty_count = 0
    path_count = 0
    prop_count = 0
    prop_total = 0

    # Only iterate over the ACTUAL data range, not formatted empty rows.
    for row in range(DATA_START_ROW, last_data_row + 1):
        # № п/п
        v = ws.cell(row=row, column=COL_NUM).value
        if v is not None and str(v).strip():
            num_count += 1

        # Кол-во
        v = ws.cell(row=row, column=COL_QTY).value
        if v is not None and str(v).strip():
            qty_count += 1

        # Путь
        v = ws.cell(row=row, column=COL_PATH).value
        if v is not None and str(v).strip():
            path_count += 1

        # Property columns
        for col in PROP_COLS:
            prop_total += 1
            v = ws.cell(row=row, column=col).value
            if v is not None and str(v).strip():
                prop_count += 1

    results["num_filled"] = num_count
    results["qty_filled"] = qty_count
    results["path_filled"] = path_count
    results["prop_filled"] = prop_count
    results["prop_total"] = prop_total

    # Check for images (embedded pictures)
    if hasattr(ws, '_images') and ws._images:
        results["has_images"] = True
    # Alternative: check _charts or drawing
    if hasattr(ws, '_charts'):
        pass

    # Validation
    if num_count == 0:
        results["issues"].append("№ п/п (col A) is EMPTY")
    elif num_count < data_rows:
        results["issues"].append(
            "№ п/п partially filled: %d/%d" % (num_count, data_rows))

    if qty_count == 0:
        results["issues"].append("Кол-во (col G) is EMPTY")
    elif qty_count < data_rows:
        results["issues"].append(
            "Кол-во partially filled: %d/%d" % (qty_count, data_rows))

    if path_count == 0:
        results["issues"].append("Путь (col O) is EMPTY")

    wb.close()
    return results


def match_mode(filename, modes):
    """Try to guess which mode produced this file based on filename or order."""
    # Files are typically named: ModelName-YYYYMMDD-HHMM.xlsx
    # We can't determine mode from filename alone.
    # User should name/number them or we process in order.
    return None


def print_report(all_results, mode_assignment=None):
    """Print validation report."""
    print("=" * 70)
    print("  ОТЧЁТ ВАЛИДАЦИИ ЭКСПОРТА BOM — 8 РЕЖИМОВ")
    print("=" * 70)
    print()

    overall_pass = True

    for i, r in enumerate(all_results):
        mode_info = ""
        if mode_assignment and i < len(mode_assignment):
            m = mode_assignment[i]
            mode_info = " → Режим %d: %s" % (m["id"], m["name"])

        status = "PASS" if not r["issues"] else "FAIL"
        if status == "FAIL":
            overall_pass = False

        print("[%s] %s%s" % (status, r["file"], mode_info))
        print("  Строк данных: %d" % r["total_rows"])
        print("  № п/п: %d/%d | Кол-во: %d/%d | Путь: %d/%d" % (
            r["num_filled"], r["total_rows"],
            r["qty_filled"], r["total_rows"],
            r["path_filled"], r["total_rows"]))
        print("  Свойства: %d/%d ячеек заполнено" % (
            r["prop_filled"], r.get("prop_total", 0)))
        if r["has_images"]:
            print("  Эскизы: ЕСТЬ")
        if r["issues"]:
            for issue in r["issues"]:
                print("  ** %s" % issue)
        print()

    # --- cross-mode sanity checks (only when files map to modes in order) ---
    if mode_assignment and len(all_results) == len(mode_assignment):
        rows_by_id = {mode_assignment[i]["id"]: all_results[i]["total_rows"]
                      for i in range(len(all_results))}
        summary = rows_by_id.get(1)
        print("ПРОВЕРКА СОГЛАСОВАННОСТИ РЕЖИМОВ:")
        if summary:
            # top-level (mode 3, type=2) must be fewer than summary
            top = rows_by_id.get(3)
            if top is not None and top >= summary:
                print("  ** Режим 3 (верхний уровень): %d строк >= сводной %d "
                      "— верхний уровень должен быть меньше." % (top, summary))
                overall_pass = False
            # filtered modes 7,8 must be a strict subset (filter applied)
            for fid, fname in ((7, "Обрабатываемые детали"),
                               (8, "Покупные изделия")):
                fr = rows_by_id.get(fid)
                if fr is None:
                    continue
                if fr == summary:
                    print("  ** Режим %d (%s): %d строк = полной сводной %d — "
                          "ФИЛЬТР НЕ ПРИМЕНЁН (полный список)."
                          % (fid, fname, fr, summary))
                    overall_pass = False
                elif fr == 0:
                    print("  ** Режим %d (%s): 0 строк — фильтр отсёк всё; "
                          "проверьте значения свойства «Тип» в модели."
                          % (fid, fname))
                else:
                    print("  - Режим %d (%s): %d из %d строк — фильтр применён."
                          % (fid, fname, fr, summary))
        print()

    print("=" * 70)
    if overall_pass:
        print("  ИТОГ: PASS — все файлы содержат данные в служебных колонках")
    else:
        print("  ИТОГ: FAIL — есть проблемы (см. выше)")
    print("=" * 70)

    # Note about property columns
    print()
    print("ПРИМЕЧАНИЕ: пустые свойства (C,D,E,H,I,J,N) — ожидаемо, если")
    print("propname в конфиге не совпадает с именами свойств модели.")
    print("На китайской демо-модели с русским конфигом свойства будут пустые.")
    print("Для полного PASS нужна модель с русскими свойствами или demo-cn конфиг.")


def main():
    if len(sys.argv) < 2:
        print("Usage: python validate_bom_exports.py <export_directory>")
        print()
        print("Export all 8 modes from ZTool, then point this script at the")
        print("directory containing the .xlsx files.")
        print()
        print("Expected: 8 .xlsx files (one per mode, in export order).")
        sys.exit(1)

    export_dir = sys.argv[1]
    if not os.path.isdir(export_dir):
        print("ERROR: '%s' is not a directory" % export_dir)
        sys.exit(1)

    # Find .xlsx files sorted by modification time (export order)
    xlsx_files = sorted(
        Path(export_dir).glob("*.xlsx"),
        key=lambda p: p.stat().st_mtime
    )

    if not xlsx_files:
        print("ERROR: no .xlsx files found in '%s'" % export_dir)
        sys.exit(1)

    print("Найдено файлов: %d" % len(xlsx_files))
    print("Ожидаемо: 8 (по одному на режим)")
    print()

    if len(xlsx_files) < 8:
        print("ВНИМАНИЕ: файлов меньше 8. Валидирую то, что есть.")
        print()

    all_results = []
    for f in xlsx_files:
        r = analyze_xlsx(str(f))
        all_results.append(r)

    # If exactly 8 files, assume they correspond to modes in order
    mode_assignment = MODES if len(xlsx_files) == 8 else None

    print_report(all_results, mode_assignment)


if __name__ == "__main__":
    main()
