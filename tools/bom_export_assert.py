#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Tier 2 acceptance: assert an exported BOM .xlsx against a golden spec.

This is the deterministic *verification* half of the semi-automated SolidWorks
acceptance (see docs/release/FULL_TEST_METHODOLOGY_RU.md, areas C/D/E). A human
or a UI-automation harness triggers the ZTool export inside SolidWorks; this
script then checks the produced workbook WITHOUT SolidWorks:

  * header set and (optionally) header ORDER match the golden;
  * data-row count matches;
  * the computed "Количество" column is numeric, > 0, and (optionally) sums to
    the expected total;
  * per-row expected cell values match;
  * per-row cell FILL color (RGB) matches the golden - this objectively replaces
    the manual "eyedropper" comparison for material/paint cell colors (COL-*).

The golden spec is captured once on a SolidWorks machine from a known-good
export (ideally the original build, for A/B parity) and committed as a fixture.

Usage:
  python tools/bom_export_assert.py <exported.xlsx> <golden.json>
  python tools/bom_export_assert.py --self-test        # exercise the asserter

Exit code 0 = all assertions passed, 1 = at least one failure.

Golden JSON schema (all keys except expected_headers optional):
  {
    "sheet": "Sheet1",                # name or 0-based index; default first sheet
    "header_row": 1,                  # 1-based row holding column headers
    "expected_headers": ["№", "Наименование", ...],
    "strict_header_order": true,      # default true
    "expected_row_count": 29,         # data rows after the header row
    "quantity_column": "Количество",  # numeric, >0 per row
    "quantity_sum": 42,               # optional total
    "rows": [                         # optional per-row assertions
      {"match": {"Обозначение": "0614-P001"},
       "expect": {"Количество": 2, "fill_rgb": "C0C0C0"}}
    ]
  }
"""
import json
import sys

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")


def norm_rgb(value):
    """Normalize an openpyxl color / hex string to upper 'RRGGBB' or None."""
    if value is None:
        return None
    # An openpyxl Color carries a real RGB only when its type is 'rgb'. For
    # 'theme'/'indexed' colors .rgb is a misleading default (e.g. '00000000'),
    # so treat anything that is not an explicit RGB color as unknown.
    ctype = getattr(value, "type", None)
    if ctype is not None and ctype != "rgb":
        return None
    rgb = getattr(value, "rgb", value)
    if rgb is None:
        return None
    if not isinstance(rgb, str):
        return None
    rgb = rgb.upper()
    if len(rgb) == 8:  # ARGB -> drop alpha
        rgb = rgb[2:]
    return rgb if len(rgb) == 6 else None


def cell_fill_rgb(cell):
    fill = cell.fill
    if fill is None or fill.patternType is None:
        return None
    return norm_rgb(fill.fgColor)


def load_sheet(path, spec):
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = spec.get("sheet")
    if sheet is None:
        ws = wb[wb.sheetnames[0]]
    elif isinstance(sheet, int):
        ws = wb[wb.sheetnames[sheet]]
    else:
        ws = wb[sheet]
    return ws


def read_table(ws, header_row):
    headers = []
    for cell in ws[header_row]:
        headers.append("" if cell.value is None else str(cell.value).strip())
    # Trim trailing empty header columns.
    while headers and headers[-1] == "":
        headers.pop()
    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        values = {}
        fills = {}
        nonempty = False
        for ci, name in enumerate(headers, start=1):
            cell = ws.cell(row=r, column=ci)
            values[name] = cell.value
            fills[name] = cell_fill_rgb(cell)
            if cell.value not in (None, ""):
                nonempty = True
        if nonempty:
            rows.append({"row": r, "values": values, "fills": fills})
    return headers, rows


class Checker:
    def __init__(self):
        self.failures = []
        self.checks = 0

    def ok(self, cond, msg):
        self.checks += 1
        if not cond:
            self.failures.append(msg)
        return cond

    def report(self):
        print("-" * 70)
        if self.failures:
            print("RESULT: FAIL - %d/%d assertion(s) failed:"
                  % (len(self.failures), self.checks))
            for f in self.failures:
                print("  * " + f)
            return 1
        print("RESULT: PASS - all %d assertion(s) passed." % self.checks)
        return 0


def to_number(value):
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return value
    try:
        return float(str(value).strip().replace(",", "."))
    except (ValueError, AttributeError):
        return None


def values_equal(got, exp):
    """Compare numerically when both sides are numbers (so 2 == 2.0), else by
    string. Real exports may serialize an integer quantity as a float."""
    gn, en = to_number(got), to_number(exp)
    if gn is not None and en is not None:
        return gn == en
    return str(got) == str(exp)


def assert_workbook(path, spec):
    c = Checker()
    ws = load_sheet(path, spec)
    header_row = int(spec.get("header_row", 1))
    headers, rows = read_table(ws, header_row)

    print("workbook:", path)
    print("sheet:", ws.title)
    print("headers:", headers)
    print("data rows:", len(rows))

    expected_headers = spec.get("expected_headers")
    if expected_headers is not None:
        if spec.get("strict_header_order", True):
            c.ok(headers == expected_headers,
                 "header order mismatch:\n      expected %s\n      got      %s"
                 % (expected_headers, headers))
        else:
            missing = [h for h in expected_headers if h not in headers]
            c.ok(not missing, "missing headers: %s" % missing)

    if "expected_row_count" in spec:
        c.ok(len(rows) == spec["expected_row_count"],
             "data-row count: expected %s, got %d"
             % (spec["expected_row_count"], len(rows)))

    qcol = spec.get("quantity_column")
    if qcol:
        if c.ok(qcol in headers, "quantity column %r not found in headers" % qcol):
            total = 0
            for rec in rows:
                n = to_number(rec["values"].get(qcol))
                if n is None or n <= 0:
                    c.ok(False, "row %d: %r must be a number > 0, got %r"
                         % (rec["row"], qcol, rec["values"].get(qcol)))
                else:
                    total += n
            if "quantity_sum" in spec:
                c.ok(total == spec["quantity_sum"],
                     "quantity sum: expected %s, got %s"
                     % (spec["quantity_sum"], total))

    for spec_row in spec.get("rows", []):
        match = spec_row.get("match", {})
        found = [r for r in rows
                 if all(str(r["values"].get(k)) == str(v)
                        for k, v in match.items())]
        if not c.ok(len(found) == 1,
                    "row match %s found %d times (expected exactly 1)"
                    % (match, len(found))):
            continue
        rec = found[0]
        for key, exp in spec_row.get("expect", {}).items():
            if key == "fill_rgb":
                # fill of the first matched column's cell, or an explicit column.
                col = spec_row.get("fill_column") or next(iter(match))
                got = rec["fills"].get(col)
                c.ok(norm_rgb(exp) == got,
                     "row %s: fill_rgb of %r expected %s, got %s"
                     % (match, col, norm_rgb(exp), got))
            else:
                got = rec["values"].get(key)
                c.ok(values_equal(got, exp),
                     "row %s: %r expected %r, got %r" % (match, key, exp, got))

    return c.report()


def self_test():
    """Build a synthetic workbook + golden in-memory and exercise pass & fail."""
    import os
    import tempfile
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill

    tmp = tempfile.mkdtemp(prefix="bom_assert_selftest_")
    xlsx = os.path.join(tmp, "export.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM"
    headers = ["№", "Наименование", "Обозначение", "Материал", "Количество"]
    ws.append(headers)
    data = [
        (1, "Кронштейн", "0614-P001", "Сталь", 2, "C0C0C0"),
        (2, "Вал", "0614-P002", "Алюминий", 1, "FFCC00"),
        (3, "Втулка", "0614-P003", "Бронза", 3, "B87333"),
    ]
    for num, name, desig, mat, qty, rgb in data:
        ws.append([num, name, desig, mat, qty])
        fill = PatternFill(start_color=rgb, end_color=rgb, fill_type="solid")
        ws.cell(row=ws.max_row, column=4).fill = fill
    wb.save(xlsx)

    good = {
        "sheet": "BOM",
        "expected_headers": headers,
        "expected_row_count": 3,
        "quantity_column": "Количество",
        "quantity_sum": 6,
        "rows": [
            {"match": {"Обозначение": "0614-P001"},
             "expect": {"Количество": 2, "fill_rgb": "C0C0C0"},
             "fill_column": "Материал"},
            {"match": {"Обозначение": "0614-P002"},
             "expect": {"fill_rgb": "FFCC00"}, "fill_column": "Материал"},
        ],
    }
    bad = dict(good)
    bad = json.loads(json.dumps(good))
    bad["expected_row_count"] = 99
    bad["rows"][0]["expect"]["fill_rgb"] = "000000"

    print("### self-test: GOOD golden (expect PASS) ###")
    rc_good = assert_workbook(xlsx, good)
    print("\n### self-test: BAD golden (expect FAIL) ###")
    rc_bad = assert_workbook(xlsx, bad)

    print("\n" + "=" * 70)
    if rc_good == 0 and rc_bad == 1:
        print("SELF-TEST PASS: asserter accepts good export and rejects bad one.")
        return 0
    print("SELF-TEST FAIL: rc_good=%d (want 0), rc_bad=%d (want 1)"
          % (rc_good, rc_bad))
    return 1


def main(argv):
    if len(argv) == 1 and argv[0] == "--self-test":
        return self_test()
    if len(argv) != 2:
        print(__doc__)
        return 2
    xlsx, golden = argv
    with open(golden, encoding="utf-8") as f:
        spec = json.load(f)
    return assert_workbook(xlsx, spec)


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
