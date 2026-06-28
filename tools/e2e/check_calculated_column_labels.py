#!/usr/bin/env python3
"""Guard calculated BOM column labels and Excel anchors.

The UI labels are deliberately short:

* Col_Weight -> "Масса"
* Col_bound  -> "Габарит"

The Excel named ranges match the visible Russian labels:

* Col_Weight -> "Масса"
* Col_bound  -> "Габарит"
"""

from __future__ import annotations

import argparse
import re
import tempfile
import textwrap
import xml.etree.ElementTree as ET
from pathlib import Path


EXPECTED = {
    "Col_Weight": ("Масса", "Масса", "J6"),
    "Col_bound": ("Габарит", "Габарит", "P6"),
}

LEGACY_VISIBLE = {
    "Col_Weight": ("Масса ед._кг", "Масса ед. кг", "Масса ед. (кг)", "МассаЕдКг"),
    "Col_bound": ("Габаритные размеры", "ГабаритныеРазмеры"),
}


class ContractError(AssertionError):
    pass


def read_text(path: Path) -> str:
    return path.read_text(encoding="utf-8-sig")


def parse_settings(path: Path) -> dict[str, tuple[str, str]]:
    root = ET.parse(path).getroot()
    result: dict[str, tuple[str, str]] = {}
    for item in root.findall(".//namemappinglist/columnnamemapping"):
        name = (item.findtext("name") or "").strip()
        if name in EXPECTED:
            result[name] = (
                (item.findtext("text") or "").strip(),
                (item.findtext("mappingname") or "").strip(),
            )
    return result


def check_settings(path: Path) -> list[str]:
    issues: list[str] = []
    mappings = parse_settings(path)
    for name, (expected_text, expected_anchor, _cell) in EXPECTED.items():
        actual = mappings.get(name)
        if actual is None:
            issues.append(f"{path}: missing mapping {name}")
            continue
        text, anchor = actual
        if text != expected_text:
            issues.append(f"{path}: {name} visible label is {text!r}, expected {expected_text!r}")
        if anchor != expected_anchor:
            issues.append(f"{path}: {name} anchor is {anchor!r}, expected {expected_anchor!r}")
    return issues


def check_source_defaults(path: Path) -> list[str]:
    source = read_text(path)
    issues: list[str] = []
    for name, (expected_text, expected_anchor, _cell) in EXPECTED.items():
        token = f'EnsureBomMapping("{name}", "{expected_text}", "{expected_anchor}")'
        if token not in source:
            issues.append(f"{path}: missing source default {token}")
        for legacy in LEGACY_VISIBLE[name]:
            if f'EnsureBomMapping("{name}", "{legacy}",' in source:
                issues.append(f"{path}: legacy default label still present for {name}: {legacy!r}")
    return issues


def check_frmmain_resx(path: Path) -> list[str]:
    source = read_text(path)
    issues: list[str] = []
    for resource_name, expected_text in (
        ("Col_Weight.HeaderText", EXPECTED["Col_Weight"][0]),
        ("Col_bound.HeaderText", EXPECTED["Col_bound"][0]),
    ):
        token = f'<data name="{resource_name}" xml:space="preserve"><value>{expected_text}</value></data>'
        if token not in source:
            issues.append(f"{path}: {resource_name} is not {expected_text!r}")
    for name, legacy_values in LEGACY_VISIBLE.items():
        for legacy in legacy_values:
            if legacy in source:
                issues.append(f"{path}: legacy visible label remains for {name}: {legacy!r}")
    return issues


def check_template(path: Path) -> list[str]:
    try:
        import openpyxl
    except Exception as exc:  # pragma: no cover - environment guard
        raise SystemExit(f"openpyxl is required: {exc}") from exc

    issues: list[str] = []
    wb = openpyxl.load_workbook(path, data_only=False)
    try:
        ws = wb.active
        for name, (expected_text, expected_anchor, cell) in EXPECTED.items():
            value = (ws[cell].value or "").strip()
            if value != expected_text:
                issues.append(f"{path}: {cell} is {value!r}, expected {expected_text!r}")
            if expected_anchor not in wb.defined_names:
                issues.append(f"{path}: missing defined name {expected_anchor!r} for {name}")
        for legacy_values in LEGACY_VISIBLE.values():
            for legacy in legacy_values:
                if any((ws[cell].value or "") == legacy for cell in ("J6", "P6")):
                    issues.append(f"{path}: legacy template header remains: {legacy!r}")
    finally:
        wb.close()
    return issues


def analyze(repo: Path) -> list[str]:
    issues: list[str] = []
    issues += check_source_defaults(repo / "client-src" / "ZTool" / "CConfigMng.cs")
    issues += check_frmmain_resx(repo / "client-src" / "ZTool.Frmmain.resx")
    issues += check_settings(repo / "SWTools.settings")
    issues += check_settings(repo / "client-core" / "dist" / "SWTools.settings")
    issues += check_template(repo / "Шаблоны спецификации" / "bom_шаблон.xlsx")
    return issues


def run_self_test() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        (root / "client-src" / "ZTool").mkdir(parents=True)
        (root / "client-core" / "dist").mkdir(parents=True)
        (root / "Шаблоны спецификации").mkdir(parents=True)
        (root / "client-src" / "ZTool" / "CConfigMng.cs").write_text(
            '\nEnsureBomMapping("Col_Weight", "Масса", "Масса");'
            '\nEnsureBomMapping("Col_bound", "Габарит", "Габарит");\n',
            encoding="utf-8",
        )
        (root / "client-src" / "ZTool.Frmmain.resx").write_text(
            '<data name="Col_Weight.HeaderText" xml:space="preserve"><value>Масса</value></data>\n'
            '<data name="Col_bound.HeaderText" xml:space="preserve"><value>Габарит</value></data>\n',
            encoding="utf-8",
        )
        settings = textwrap.dedent(
            """\
            <CConfigDO>
              <namemappinglist>
                <columnnamemapping><name>Col_Weight</name><text>Масса</text><mappingname>Масса</mappingname></columnnamemapping>
                <columnnamemapping><name>Col_bound</name><text>Габарит</text><mappingname>Габарит</mappingname></columnnamemapping>
              </namemappinglist>
            </CConfigDO>
            """
        )
        (root / "SWTools.settings").write_text(settings, encoding="utf-8")
        (root / "client-core" / "dist" / "SWTools.settings").write_text(settings, encoding="utf-8")

        import openpyxl
        from openpyxl.workbook.defined_name import DefinedName

        wb = openpyxl.Workbook()
        ws = wb.active
        ws["J6"] = "Масса"
        ws["P6"] = "Габарит"
        wb.defined_names.add(DefinedName("Масса", attr_text=f"'{ws.title}'!$J$6"))
        wb.defined_names.add(DefinedName("Габарит", attr_text=f"'{ws.title}'!$P$6"))
        wb.save(root / "Шаблоны спецификации" / "bom_шаблон.xlsx")
        wb.close()

        good = analyze(root)
        if good:
            raise ContractError(f"good fixture failed: {good}")

        bad_resx = root / "client-src" / "ZTool.Frmmain.resx"
        bad_resx.write_text(read_text(bad_resx).replace("Габарит", "Габаритные размеры"), encoding="utf-8")
        bad = analyze(root)
        if not any("legacy visible label remains" in issue for issue in bad):
            raise ContractError(f"bad fixture unexpectedly passed: {bad}")
    print("self-test PASS")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--repo-root", default=str(Path(__file__).resolve().parents[2]))
    parser.add_argument("--self-test", action="store_true")
    args = parser.parse_args()

    if args.self_test:
        run_self_test()
        return 0

    issues = analyze(Path(args.repo_root))
    if issues:
        for issue in issues:
            print(f"FAIL: {issue}")
        return 1
    print("PASS: calculated BOM columns use consistent labels/anchors Масса/Габарит")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
